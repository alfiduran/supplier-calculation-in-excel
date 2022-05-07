import openpyxl #you will have to pip install this package. try "pip3 install openpyxl" if pip doesnt work

inv_file = openpyxl.load_workbook("inventory.xlsx")
product_list = inv_file["Sheet1"] #which sheet you looking at?


# Task: calculate how many products we have per supplier and list the supplier names
product_per_supplier = {}

# print(product_list.max_row)

# looping! How many times we wanna loops? here we got 74 rows but we want it to run as many times
# we need to think about the logic. 
for product_row in range(2, product_list.max_row + 1): #check documentations +1 is to addd the last line. because exclusivity
    supplier_name = product_list.cell(product_row, 4).value
    
    if supplier_name in product_per_supplier:
        current_num_products = product_per_supplier[supplier_name]
        product_per_supplier[supplier_name] = current_num_products + 1
    else:
        print("Adding new supplier")
        product_per_supplier[supplier_name] = 1

print(product_per_supplier)
